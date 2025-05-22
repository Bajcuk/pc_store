from PySide6.QtWidgets import QMainWindow, QMessageBox, QTableWidgetItem, QTableWidget, QPushButton, QHBoxLayout, \
    QDialog, QFileDialog
from PySide6.QtCore import Qt

import openpyxl
from openpyxl.utils import get_column_letter


from app.views.ui_edit_component import EditComponentDialog
from app.views.ui_main_window import Ui_MainWindow
from app.models.database import (
    get_components_with_category_name,
    get_all_categories, get_users_with_access_level,
    AccessLevel, add_component, get_all_components,
    update_component, delete_component
)

class MainWindow(QMainWindow):
    def __init__(self, login_window, user_data):
        super().__init__()

        if user_data['access_level'] == AccessLevel.UNVERIFIED:
            self.close()
            login_window.show_unverified_window()
            return

        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.login_window = login_window
        self.user_data = user_data
        self.setWindowTitle(f"Управление ({user_data['name']})")

        self.current_category_id = None
        self.current_search_text = ""

        self.load_components()
        self.load_categories()

        if self.user_data['access_level'] == AccessLevel.ADMIN:
            self.load_users()
            self.ui.tab_widget.setTabEnabled(1, True)
        else:
            self.ui.tab_widget.setTabEnabled(1, False)

        self.setup_connections()

    def setup_connections(self):
        self.ui.category_filter.currentIndexChanged.connect(self.on_category_changed)
        self.ui.search_field.textChanged.connect(self.on_search_changed)

        self.ui.button_refresh.clicked.connect(self.refresh_data)
        self.ui.button_logout.clicked.connect(self.logout)

        self.ui.button_add.clicked.connect(self.add_component)
        self.ui.button_edit.clicked.connect(self.edit_component)
        self.ui.button_delete.clicked.connect(self.delete_component)

        self.ui.btn_update_access.clicked.connect(self.update_user_access)
        self.ui.btn_delete_user.clicked.connect(self.delete_user)

        self.ui.tab_widget.currentChanged.connect(self.on_tab_changed)

        self.ui.button_export.clicked.connect(self.export_to_excel)

    def load_categories(self):
        try:
            self.ui.category_filter.clear()
            self.ui.category_filter.addItem("Все категории", None)
            categories = get_all_categories()
            for category in categories:
                self.ui.category_filter.addItem(
                    f"{category.name} ({category.description})", category.category_id
                )
            self.current_category_id = None
            self.ui.category_filter.setCurrentIndex(0)
        except Exception as e:
            print(f"Ошибка загрузки категорий: {e}")
            QMessageBox.warning(self, "Ошибка", "Не удалось загрузить список категорий")

    def load_components(self):
        components = get_components_with_category_name()
        self.ui.table_components.setColumnCount(6)
        self.ui.table_components.setHorizontalHeaderLabels([
            "ID", "Название", "Описание", "Количество", "Цена", "Категория"
        ])
        self.ui.table_components.setRowCount(len(components))
        for row_idx, component in enumerate(components):
            self.ui.table_components.setItem(row_idx, 0, QTableWidgetItem(str(component.component_id)))
            self.ui.table_components.setItem(row_idx, 1, QTableWidgetItem(component.name))
            self.ui.table_components.setItem(row_idx, 2, QTableWidgetItem(component.description))
            self.ui.table_components.setItem(row_idx, 3, QTableWidgetItem(str(component.quantity)))
            self.ui.table_components.setItem(row_idx, 4, QTableWidgetItem(f"{component.price:.2f} ₽"))
            self.ui.table_components.setItem(row_idx, 5, QTableWidgetItem(component.category_name))
            for col in [0, 3, 4]:
                item = self.ui.table_components.item(row_idx, col)
                item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.ui.table_components.resizeColumnsToContents()

    def load_users(self):
        users = get_users_with_access_level()
        self.ui.table_users.setRowCount(len(users))
        self.ui.table_users.setColumnCount(5)
        self.ui.table_users.setHorizontalHeaderLabels([
            "ID", "Имя", "Фамилия", "Логин", "Уровень доступа"
        ])
        for row_idx, user in enumerate(users):
            self.ui.table_users.setItem(row_idx, 0, QTableWidgetItem(str(user['user_id'])))
            self.ui.table_users.setItem(row_idx, 1, QTableWidgetItem(user['name']))
            self.ui.table_users.setItem(row_idx, 2, QTableWidgetItem(user['last_name']))
            self.ui.table_users.setItem(row_idx, 3, QTableWidgetItem(user['login']))
            self.ui.table_users.setItem(row_idx, 4, QTableWidgetItem(user['access_level']))
            self.ui.table_users.item(row_idx, 0).setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.ui.table_users.resizeColumnsToContents()

    def on_category_changed(self, index):
        self.current_category_id = self.ui.category_filter.currentData()
        self.apply_filters()

    def on_search_changed(self, text):
        self.current_search_text = text.strip().lower()
        self.apply_filters()

    def apply_filters(self):
        try:
            all_components = get_components_with_category_name()
            filtered = []
            for comp in all_components:
                if (self.current_category_id is None or comp.category_id == self.current_category_id) and \
                   (not self.current_search_text or self.current_search_text in comp.name.lower()):
                    filtered.append(comp)
            self.update_components_table(filtered)
        except Exception as e:
            print(f"Ошибка фильтрации: {e}")
            QMessageBox.warning(self, "Ошибка", "Не удалось применить фильтры")

    def update_components_table(self, components):
        self.ui.table_components.setRowCount(len(components))
        for row_idx, comp in enumerate(components):
            self.ui.table_components.setItem(row_idx, 0, QTableWidgetItem(str(comp.component_id)))
            self.ui.table_components.setItem(row_idx, 1, QTableWidgetItem(comp.name))
            self.ui.table_components.setItem(row_idx, 2, QTableWidgetItem(comp.description))
            self.ui.table_components.setItem(row_idx, 3, QTableWidgetItem(str(comp.quantity)))
            self.ui.table_components.setItem(row_idx, 4, QTableWidgetItem(f"{comp.price:.2f} ₽"))
            self.ui.table_components.setItem(row_idx, 5, QTableWidgetItem(comp.category_name))
            for col in [0, 3, 4]:
                item = self.ui.table_components.item(row_idx, col)
                item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.ui.table_components.resizeColumnsToContents()

    def refresh_data(self):
        self.current_category_id = None
        self.current_search_text = ""
        self.ui.search_field.clear()
        self.ui.category_filter.setCurrentIndex(0)
        self.load_categories()
        self.load_components()
        if self.user_data['access_level'] == AccessLevel.ADMIN:
            self.load_users()

    def add_component(self):
        categories = get_all_categories()
        dialog = EditComponentDialog(categories, self)
        if dialog.exec() == QDialog.Accepted:
            data = dialog.get_data()
            if add_component(**data):
                self.refresh_data()
                QMessageBox.information(self, "Успех", "Товар успешно добавлен")
            else:
                QMessageBox.warning(self, "Ошибка", "Не удалось добавить товар")

    def edit_component(self):
        selected = self.ui.table_components.selectedItems()
        if not selected:
            QMessageBox.warning(self, "Ошибка", "Выберите товар")
            return

        component_id = int(self.ui.table_components.item(selected[0].row(), 0).text())
        components = get_all_components()
        component = next((c for c in components if c.component_id == component_id), None)

        if not component:
            QMessageBox.warning(self, "Ошибка", "Товар не найден")
            return

        categories = get_all_categories()
        dialog = EditComponentDialog(categories, self, component)
        if dialog.exec() == QDialog.Accepted:
            data = dialog.get_data()
            if update_component(component_id, **data):
                self.refresh_data()
                QMessageBox.information(self, "Успех", "Товар обновлён")
            else:
                QMessageBox.warning(self, "Ошибка", "Не удалось обновить товар")

    def delete_component(self):
        selected = self.ui.table_components.selectedItems()
        if not selected:
            QMessageBox.warning(self, "Ошибка", "Выберите товар")
            return

        component_id = int(self.ui.table_components.item(selected[0].row(), 0).text())
        reply = QMessageBox.question(self, "Подтверждение", "Удалить товар?",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            if delete_component(component_id):
                self.refresh_data()
                QMessageBox.information(self, "Успех", "Товар удалён")
            else:
                QMessageBox.warning(self, "Ошибка", "Не удалось удалить")

    def update_user_access(self):
        selected = self.ui.table_users.selectedItems()
        if not selected:
            QMessageBox.warning(self, "Ошибка", "Выберите пользователя")
            return

        user_id = int(self.ui.table_users.item(selected[0].row(), 0).text())
        user_login = self.ui.table_users.item(selected[0].row(), 3).text()
        new_level = self.ui.combo_access_level.currentData()

        if user_login == self.user_data['login']:
            QMessageBox.warning(self, "Ошибка", "Нельзя менять свои права")
            return

        reply = QMessageBox.question(self, "Подтверждение",
                                     f"Изменить права {user_login} на '{new_level}'?",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            from app.models.database import update_user_access
            if update_user_access(user_id, new_level):
                QMessageBox.information(self, "Успех", "Права обновлены")
                self.load_users()
            else:
                QMessageBox.warning(self, "Ошибка", "Не удалось обновить")

    def delete_user(self):
        selected = self.ui.table_users.selectedItems()
        if not selected:
            QMessageBox.warning(self, "Ошибка", "Выберите пользователя")
            return

        user_id = int(self.ui.table_users.item(selected[0].row(), 0).text())
        user_login = self.ui.table_users.item(selected[0].row(), 3).text()

        if user_login == self.user_data['login']:
            QMessageBox.warning(self, "Ошибка", "Нельзя удалить себя")
            return

        reply = QMessageBox.question(self, "Подтверждение",
                                     f"Удалить пользователя {user_login}?",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            from app.models.database import delete_user
            if delete_user(user_id):
                QMessageBox.information(self, "Успех", "Пользователь удалён")
                self.load_users()
            else:
                QMessageBox.warning(self, "Ошибка", "Не удалось удалить")

    def on_tab_changed(self, index):
        if self.ui.tab_widget.tabText(index) == "Комплектующие":
            self.ui.button_add.show()
            self.ui.button_edit.show()
            self.ui.button_delete.show()
            self.ui.button_export.show()
        else:
            self.ui.button_add.hide()
            self.ui.button_edit.hide()
            self.ui.button_delete.hide()
            self.ui.button_export.hide()

    def export_to_excel(self):
        path, _ = QFileDialog.getSaveFileName(self, "Сохранить как Excel", "", "Excel Files (*.xlsx)")
        if not path:
            return

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Комплектующие"

        # Заголовки
        headers = ["ID", "Название", "Описание", "Количество", "Цена", "Категория"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)

        # Данные из таблицы
        for row in range(self.ui.table_components.rowCount()):
            for col in range(self.ui.table_components.columnCount()):
                item = self.ui.table_components.item(row, col)
                if item:
                    sheet.cell(row=row + 2, column=col + 1, value=item.text())

        # Автоширина
        for col in range(1, self.ui.table_components.columnCount() + 1):
            sheet.column_dimensions[get_column_letter(col)].auto_size = True

        try:
            workbook.save(path)
            QMessageBox.information(self, "Успех", "Экспорт завершён")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить файл: {e}")

    def logout(self):
        self.close()
        self.login_window.show()
